import tkinter as tk
import tkinter.filedialog
import random
import openpyxl


class TeamManager:
    """A class for managing a list of teams."""

    def __init__(self):
        self.teams = []
        self.counter = 1

    def add_team(self, name):
        """Add a team to the list of teams."""
        self.teams.append((self.counter, name))
        self.counter += 1

    def load_teams(self, filename):
        """Load teams from a file."""
        try:
            with open(filename, "r", encoding='utf8') as file:
                for line in file:
                    self.add_team(line.strip())
        except FileNotFoundError:
            tk.messagebox.showerror("Error", "File not found.")
        except Exception as e:
            tk.messagebox.showerror("Error", f"Error reading file: {e}")

    def select_teams(self):
        """Select two teams at random from the list of teams."""
        if len(self.teams) >= 2:
            selected = random.sample(self.teams, 2)
            self.teams.remove(selected[0])
            self.teams.remove(selected[1])
            return selected
        else:
            return None

    def clear_teams(self):
        """Clear the list of teams. Are you sure you want to clear the teams list?"""
        if tk.messagebox.askyesno("Confirmation", "Вы действительно хотите очистить список команд?"):
            self.teams = []
            self.counter = 1


team_manager = TeamManager()


class TeamManagerGUI(tk.Frame):
    """A GUI for the team manager."""

    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.master.title("Team Manager")
        self.pack(fill=tk.BOTH, expand=True)
        self.create_widgets()

    # def create_widgets(self):
    #     self.load_button = tk.Button(self, text="Загрузить команды", font=("Cascadia Code ExtraLight", 14),
    #                                  command=self.load_teams)
    #     self.load_button.grid(row=0, column=0, padx=10, pady=10)
    #     self.select_teams_button = tk.Button(self, text="Выбрать команды", font=("Cascadia Code ExtraLight", 14),
    #                                          command=self.select_teams)
    #     self.select_teams_button.grid(row=0, column=1, padx=10, pady=10)
    #     self.clear_teams_button = tk.Button(self, text="Очистить список команд", font=("Cascadia Code ExtraLight", 14),
    #                                         command=self.clear_teams)
    #     self.clear_teams_button.grid(row=0, column=2, padx=10, pady=10)
    #     self.quit_button = tk.Button(self, text="Выход", font=("Cascadia Code ExtraLight", 14),
    #                                  command=self.master.destroy)
    #     self.quit_button.grid(row=0, column=3, padx=10, pady=10)
    #     self.teams_listbox = tk.Listbox(self, font=("Cascadia Code ExtraLight", 12))
    #     self.teams_listbox.grid(row=1, column=0, columnspan=3, padx=10, pady=10, sticky=tk.NSEW)
    #
    #     # self.font_sizebox = tk.Listbox(self, font=("Cascadia Code ExtraLight", 12))
    #     # self.font_sizebox.grid(row=1, column=3, columnspan=1, padx=10, pady=10)
    #
    #     self.selected_teams_var = tk.StringVar()
    #     self.selected_teams_label = tk.Label(self, textvariable=self.selected_teams_var,
    #                                          font=("Cascadia Code ExtraLight", 35))
    #     self.selected_teams_label.grid(row=2, column=0, columnspan=4, padx=10, pady=10, sticky=tk.NSEW)

    def create_widgets(self):
        self.load_button = tk.Button(self, text="Загрузить команды", font=("Cascadia Code ExtraLight", 14),
                                     command=self.load_teams)
        self.load_button.grid(row=0, column=0, padx=10, pady=10)
        self.select_teams_button = tk.Button(self, text="Выбрать команды", font=("Cascadia Code ExtraLight", 14),
                                             command=self.select_teams)
        self.select_teams_button.grid(row=0, column=1, padx=10, pady=10)
        self.clear_teams_button = tk.Button(self, text="Очистить список команд", font=("Cascadia Code ExtraLight", 14),
                                            command=self.clear_teams)
        self.clear_teams_button.grid(row=0, column=2, padx=10, pady=10)
        self.quit_button = tk.Button(self, text="Выход", font=("Cascadia Code ExtraLight", 14),
                                     command=self.master.destroy)
        self.quit_button.grid(row=0, column=3, padx=10, pady=10)
        self.teams_listbox = tk.Listbox(self, font=("Cascadia Code ExtraLight", 12))
        self.teams_listbox.grid(row=1, column=0, columnspan=3, padx=10, pady=10, sticky=tk.NSEW)
        self.font_size_button = tk.Button(self, text="Размер шрифта", font=("Cascadia Code ExtraLight", 14),
                                          command=self.change_font_size)
        self.font_size_button.grid(row=1, column=3, padx=10, pady=10)
        self.selected_teams_var = tk.StringVar()
        self.selected_teams_label = tk.Label(self, textvariable=self.selected_teams_var,
                                             font=("Cascadia Code ExtraLight", 35))
        self.selected_teams_label.grid(row=2, column=0, columnspan=4, padx=10, pady=10, sticky=tk.NSEW)

        # Assign weights to rows and columns
        for i in range(4):
            self.grid_columnconfigure(i, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=2)
        self.grid_rowconfigure(2, weight=2)

    def change_font_size(self):
        font_size_window = tk.Toplevel(self)
        font_size_window.title("Размер шрифта")
        font_size_label = tk.Label(font_size_window, text="Введите размер шрифта:")
        font_size_label.pack()
        font_size_entry = tk.Entry(font_size_window)
        font_size_entry.pack()
        ok_button = tk.Button(font_size_window, text="OK", command=lambda: self.apply_font_size(font_size_entry.get()))
        ok_button.pack()

    def apply_font_size(self, font_size):
        self.selected_teams_label.config(font=("Cascadia Code ExtraLight", font_size))

    def load_teams(self):
        filename = tk.filedialog.askopenfilename()
        if filename:
            team_manager.load_teams(filename)
            self.teams_listbox.delete(0, tk.END)
            for team in team_manager.teams:
                self.teams_listbox.insert(tk.END, f"{team[0]}: {team[1]}")

    def select_teams(self):
        selected = team_manager.select_teams()
        if selected:
            self.selected_teams_var.set(
                f"{selected[0][0]} - {selected[0][1]}\n vs \n{selected[1][0]} - {selected[1][1]}")
            self.teams_listbox.delete(0, tk.END)
            for team in team_manager.teams:
                self.teams_listbox.insert(tk.END, f"{team[0]}: {team[1]}")
            # Add selected teams to Excel file
            wb = openpyxl.load_workbook("teams.xlsx")
            ws = wb.active
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=selected[0][1])
            ws.cell(row=row, column=2, value=selected[1][1])
            wb.save("teams.xlsx")

    def clear_teams(self):
        team_manager.clear_teams()
        self.teams_listbox.delete(0, tk.END)
        self.selected_teams_var.set("")


root = tk.Tk()
root.geometry("1500x950")
app = TeamManagerGUI(master=root)
app.mainloop()
